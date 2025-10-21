import io
import re
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from psycopg2 import sql

from views.config import (
    AKUN_REGULAR,
    MARKETPLACE_LIST,
    PLATFORM_REGULAR,
    PROJECT_NAME_LIST,
    TOKO_BANDUNG,
    get_yesterday_in_jakarta,
)


# -- ORDERS DATA
def clean_admin_marketplace_data(df: pd.DataFrame) -> pd.DataFrame:
    """
    Fungsi untuk membersihkan dan memproses data pesanan admin marketplace dari BigSeller.

    Args:
        df (pd.DataFrame): DataFrame mentah dari file yang diunggah.

    Returns:
        pd.DataFrame: DataFrame yang sudah bersih dan terstandarisasi.
    """
    df = df.copy()

    columns_mapping = {
        "Nomor Pesanan": "order_id",
        "Nomor Paket": "package_id",
        "Status Pesanan": "order_status",
        "Marketplace": "nama_marketplace",
        "Toko Marketplace": "nama_toko",
        "Nama Pembeli": "nama_pembeli",
        "Nomor Telepon": "no_telepon",
        "Kode Pos": "kode_pos",
        "Negara": "negara",
        "Provinsi": "provinsi",
        "Kabupaten/Kota": "kabupaten_kota",
        "Kecamatan": "kecamatan",
        "Kelurahan": "kelurahan",
        "Alamat Lengkap": "alamat_lengkap",
        "SKU": "sku",
        "Nama Produk": "nama_produk",
        "Jumlah": "jumlah",
        "Harga Satuan": "harga_satuan",
        "Subtotal Produk": "subtotal_produk",
        "Harga Awal Produk": "harga_awal_produk",
        "Gudang Asal": "gudang_asal",
        "Jasa Kirim yang Dipilih Pembeli": "jasa_kirim",
        "Nomor Resi": "no_resi",
        "Ongkos Kirim": "ongkos_kirim",
        "Diskon Ongkos Kirim Penjual": "diskon_ongkos_kirim_penjual",
        "Diskon Ongkos Kirim Marketplace": "diskon_ongkos_kirim_marketplace",
        "Total Pesanan": "total_pesanan",
        "Metode Pembayaran": "metode_pembayaran",
        "Biaya Pengelolaan": "biaya_pengelolaan",
        "Biaya Transaksi": "biaya_transaksi",
        "Diskon Penjual": "diskon_penjual",
        "Diskon Marketplace": "diskon_marketplace",
        "Voucher": "voucher",
        "Voucher Toko": "voucher_toko",
        "Waktu Pesanan Dibuat": "waktu_pesanan_dibuat",
        "Waktu Pesanan Dibayar": "waktu_pesanan_dibayar",
        "Waktu Kedaluwarsa": "waktu_kadaluwarsa",
        "Waktu Proses": "waktu_proses",
        "Waktu Cetak": "waktu_cetak",
        "Waktu Pesanan Dikirim": "waktu_pesanan_dikirim",
        "Waktu Selesai": "waktu_selesai",
        "Waktu Pembatalan": "waktu_pembatalan",
        "Pesan dari Pembeli": "pesan_dari_pembeli",
        "Yang Membatalkan": "yang_membatalkan",
    }

    # Ganti nama kolom dan ambil kolom yang relevan
    df.rename(columns=columns_mapping, inplace=True)
    df = df[list(columns_mapping.values())]

    object_cols = df.select_dtypes(["object"]).columns.to_list()
    for col in object_cols:
        df[col] = df[col].str.strip()
        df[col] = df[col].str.replace(r"\xa0", " ", regex=True)

    # Standarisasi kolom teks (mengubah ke title case dan lowercase)
    df["nama_toko"] = df["nama_toko"].str.lower()
    df["sku"] = df["sku"].str.replace(r"\r\n", " ", regex=True)

    address_cols = [
        "negara",
        "provinsi",
        "kabupaten_kota",
        "kecamatan",
        "kelurahan",
        "alamat_lengkap",
        "nama_produk",
        "nama_pembeli",
    ]
    for col in address_cols:
        df[col] = df[col].str.lower()

    marketplace_map = {
        "Shopee": "SP",
        "Lazada": "LZ",
        "Tokopedia": "TP",
        "TikTok": "TT",
    }

    df["nama_toko"] = (
        df["nama_marketplace"].map(marketplace_map) + " " + df["nama_toko"].str.lower()
    )

    # df["marketplace_toko"] = df["nama_marketplace"] + df["nama_toko"]
    # df["nama_toko_standard"] = df["marketplace_toko"].map(store_mapping)

    # Konversi kolom tanggal dan waktu ke tipe datetime
    month_mapping = {
        "Jan": "Jan",
        "Feb": "Feb",
        "Mar": "Mar",
        "Apr": "Apr",
        "Mei": "May",
        "Jun": "Jun",
        "Jul": "Jul",
        "Agu": "Aug",
        "Sep": "Sep",
        "Okt": "Oct",
        "Nov": "Nov",
        "Des": "Dec",
    }

    # compile regex untuk ganti bulan (lebih cepat daripada loop biasa)
    pattern = re.compile(r"\b(" + "|".join(month_mapping.keys()) + r")\b")

    def replace_months(series: pd.Series) -> pd.Series:
        """Ganti bulan Indonesia ke format Inggris + parsing datetime."""
        replaced = series.astype(str).str.replace(
            pattern, lambda m: month_mapping.get(m.group()), regex=True
        )

        return pd.to_datetime(replaced, errors="coerce")

    datetime_cols = [
        "waktu_pesanan_dibuat",
        "waktu_pesanan_dibayar",
        "waktu_kadaluwarsa",
        "waktu_proses",
        "waktu_cetak",
        "waktu_pesanan_dikirim",
        "waktu_selesai",
        "waktu_pembatalan",
    ]

    df[datetime_cols] = df[datetime_cols].apply(replace_months)

    # Konversi kolom numerik
    currency_cols = [
        "harga_satuan",
        "subtotal_produk",
        "harga_awal_produk",
        "ongkos_kirim",
        "diskon_ongkos_kirim_penjual",
        "diskon_ongkos_kirim_marketplace",
        "total_pesanan",
        "biaya_pengelolaan",
        "biaya_transaksi",
        "diskon_penjual",
        "diskon_marketplace",
        "voucher",
        "voucher_toko",
    ]
    for col in currency_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Standarisasi kolom numerik yang mungkin perlu konversi ke float
    # df["ongkos_kirim"] = df["ongkos_kirim"].astype(float)

    # df.drop(columns="marketplace_toko", inplace=True)

    return df


def add_new_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Fungsi untuk menambahkan kolom-kolom baru berdasarkan logika bisnis.

    Args:
        df (pd.DataFrame): DataFrame yang sudah bersih.

    Returns:
        pd.DataFrame: DataFrame dengan kolom-kolom baru.
    """
    df = df.copy()

    # Tambah kolom baru
    df["is_fake_order"] = np.where(
        df["pesan_dari_pembeli"].str.contains("FO", na=False), True, False
    )

    df["gudang_asal"] = np.where(
        df["nama_toko"].isin(TOKO_BANDUNG), "Bandung", "Jakarta"
    )

    brand_map = {
        "ZYY": "Zhi Yang Yao",
        "ERA": "Erassgo",
        "ENZ": "Enzhico",
        "KDK": "Kudaku",
        "JOY": "Joymens",
        "GLU": "Glumevit",
        "GOD": "Godfather",
        "BTN": "Bycetin",
        "BIO": "Bio Antanam",
        "DOU": "Doui",
    }

    # ambil prefix SKU sebelum tanda '-'
    df["nama_brand"] = df["sku"].str.split("-", n=1).str[0].map(brand_map)
    df["nama_brand"].fillna("Unknown", inplace=True)

    return df


def add_timestamp_and_sesi_columns(
    df: pd.DataFrame, timestamp_input_data: datetime
) -> pd.DataFrame:
    """
    Fungsi untuk menambahkan kolom timestamp_input_data dan sesi ke DataFrame.

    Args:
        df (pd.DataFrame): DataFrame yang sudah dibersihkan.
        timestamp_input_data (datetime): Objek datetime yang berisi tanggal dan waktu input data.

    Returns:
        pd.DataFrame: DataFrame dengan kolom tambahan.
    """
    df["timestamp_input_data"] = timestamp_input_data

    hours = timestamp_input_data.hour
    conditions = [
        (hours >= 7) & (hours < 12),
        (hours >= 12) & (hours < 14),
        (hours >= 14) & (hours < 16),
    ]
    choices = ["SESI 1", "SESI 2", "SESI 3"]
    df["sesi"] = np.select(conditions, choices, default="SESI 4")

    return df


# FINANCE DATA
def initialize_omset_data_session(project_name, marketplace_list, store_list):
    """
    Menginisialisasi DataFrame di st.session_state untuk brand tertentu jika belum ada.

    Args:
        project_name (str): Nama project untuk kunci di session_state.
        marketplace__list (list): Daftar nama marketplace yang akan diisi ke DataFrame.
        store_list (list): Daftar nama toko yang akan diisi ke DataFrame.
    """
    session_key = f"df_{project_name}_omset"

    if session_key not in st.session_state:
        # Buat DataFrame default dengan kolom yang dibutuhkan
        data = {
            "Tanggal": pd.Series(
                get_yesterday_in_jakarta(),
                index=range(len(store_list)),
            ),
            "Marketplace": marketplace_list,
            "Nama Toko": store_list,
            # "Akrual Basis": [0.0] * len(store_list),
            "Pendapatan Kotor": [0.0] * len(store_list),
            "Biaya Admin": [0.0] * len(store_list),
            "Cash Basis": [0.0] * len(store_list),
            "Bukti": [None] * len(store_list),
            "Akun Bank": [None] * len(store_list),
        }
        st.session_state[session_key] = pd.DataFrame(data)


def get_omset_column_config(store_list):
    """
    Mengembalikan konfigurasi kolom yang konsisten untuk st.data_editor.
    """

    return {
        "Tanggal": st.column_config.DateColumn(
            "Tanggal",
            min_value=pd.Timestamp(2023, 1, 1),
            format="YYYY-MM-DD",
            required=True,
        ),
        "Marketplace": st.column_config.SelectboxColumn(
            "Marketplace",
            options=MARKETPLACE_LIST,
            required=True,
        ),
        "Nama Toko": st.column_config.SelectboxColumn(
            "Nama Toko",
            options=store_list,
            required=True,
        ),
        # "Akrual Basis": st.column_config.NumberColumn(
        #     "Akrual Basis (Rp)",
        #     min_value=0.0,
        #     format="accounting",
        #     required=True,
        # ),
        "Pendapatan Kotor": st.column_config.NumberColumn(
            "Pendapatan Kotor (Rp)",
            min_value=0.0,
            format="accounting",
            required=True,
        ),
        "Biaya Admin": st.column_config.NumberColumn(
            "Biaya Admin (Rp)",
            min_value=0.0,
            format="accounting",
            required=True,
        ),
        "Cash Basis": st.column_config.NumberColumn(
            "Cash Basis (Rp)",
            min_value=0.0,
            format="accounting",
        ),
        "Bukti": st.column_config.TextColumn("Bukti"),
        "Akun Bank": st.column_config.TextColumn("Akun Bank"),
    }


# FINANCE DATA
def initialize_omset_reg_data_session(project_name):
    """
    Menginisialisasi DataFrame di st.session_state untuk brand tertentu jika belum ada.
    """
    session_key = f"df_{project_name}_omset_reg"

    if session_key not in st.session_state:
        # Buat DataFrame default dengan kolom yang dibutuhkan
        data = {
            "Tanggal": pd.Series(
                get_yesterday_in_jakarta(),
                index=range(len(PLATFORM_REGULAR)),
            ),
            "Platform": PLATFORM_REGULAR,
            "Akrual Basis": [0.0] * len(PLATFORM_REGULAR),
            "Cash Basis": [0.0] * len(PLATFORM_REGULAR),
            "Bukti": [None] * len(PLATFORM_REGULAR),
            "Akun Bank": [None] * len(PLATFORM_REGULAR),
        }
        st.session_state[session_key] = pd.DataFrame(data)


def get_omset_reg_column_config():
    """
    Mengembalikan konfigurasi kolom yang konsisten untuk st.data_editor.
    """

    return {
        "Tanggal": st.column_config.DateColumn(
            "Tanggal",
            min_value=pd.Timestamp(2023, 1, 1),
            format="YYYY-MM-DD",
            required=True,
        ),
        "Platform": st.column_config.SelectboxColumn(
            "Platform",
            options=PLATFORM_REGULAR,
            required=True,
        ),
        "Akrual Basis": st.column_config.NumberColumn(
            "Akrual Basis (Rp)",
            min_value=0.0,
            format="accounting",
            required=True,
        ),
        "Cash Basis": st.column_config.NumberColumn(
            "Cash Basis (Rp)",
            min_value=0.0,
            format="accounting",
        ),
        "Bukti": st.column_config.TextColumn("Bukti"),
        "Akun Bank": st.column_config.TextColumn("Akun Bank"),
    }


def initialize_ads_data_session(project_name, marketplace_list, store_list):
    """
    Menginisialisasi DataFrame di st.session_state untuk project tertentu jika belum ada.

    Args:
        project_name (str): Nama project untuk kunci di session_state.
        marketplace__list (list): Daftar nama marketplace yang akan diisi ke DataFrame.
        store_list (list): Daftar nama toko yang akan diisi ke DataFrame.
    """
    session_key = f"df_{project_name}_ads"

    if session_key not in st.session_state:
        # Buat DataFrame default dengan kolom yang dibutuhkan
        data = {
            "Tanggal": pd.Series(
                get_yesterday_in_jakarta(),
                index=range(len(store_list)),
            ),
            "Marketplace": marketplace_list,
            "Nama Toko": store_list,
            # "Nominal Budget Ads": [0.0] * len(store_list),
            "Nominal Aktual Ads": None,
        }
        st.session_state[session_key] = pd.DataFrame(data)


def get_ads_column_config(store_list):
    """
    Mengembalikan konfigurasi kolom yang konsisten untuk st.data_editor.
    """

    return {
        "Tanggal": st.column_config.DateColumn(
            "Tanggal",
            min_value=pd.Timestamp(2023, 1, 1),
            format="YYYY-MM-DD",
            required=True,
        ),
        "Marketplace": st.column_config.SelectboxColumn(
            "Marketplace",
            options=MARKETPLACE_LIST,
            required=True,
        ),
        "Nama Toko": st.column_config.SelectboxColumn(
            "Nama Toko",
            options=store_list,
            required=True,
        ),
        "Nominal Aktual Ads": st.column_config.NumberColumn(
            "Nominal Aktual Ads (Rp)",
            min_value=0.0,
            format="accounting",
        ),
    }


def initialize_finance_cpas_data_session(project_name, store_list, akun_list):
    """
    Menginisialisasi DataFrame di st.session_state untuk store_list tertentu jika belum ada.

    Args:
        project_name (str): Nama store_list untuk kunci di session_state.
        marketplace__list (list): Daftar nama marketplace yang akan diisi ke DataFrame.
        akun (list): Daftar nama toko yang akan diisi ke DataFrame.
    """
    session_key = f"df_{project_name}_cpas"

    if session_key not in st.session_state:
        # Buat DataFrame default dengan kolom yang dibutuhkan
        data = {
            "Tanggal": pd.Series(
                get_yesterday_in_jakarta(),
                index=range(len(akun_list)),
            ),
            "Nama Toko": store_list,
            "Akun": akun_list,
            "Nominal Aktual Ads": [0.0] * len(akun_list),
        }
        st.session_state[session_key] = pd.DataFrame(data)


def get_finance_cpas_column_config(store_list, akun_list):
    """
    Mengembalikan konfigurasi kolom yang konsisten untuk st.data_editor.
    """

    return {
        "Tanggal": st.column_config.DateColumn(
            "Tanggal",
            min_value=pd.Timestamp(2023, 1, 1),
            format="YYYY-MM-DD",
            required=True,
        ),
        "Nama Toko": st.column_config.SelectboxColumn(
            "Nama Toko",
            options=store_list,
            required=True,
        ),
        "Akun": st.column_config.SelectboxColumn(
            "Akun",
            options=akun_list,
            required=True,
        ),
        "Nominal Aktual Ads": st.column_config.NumberColumn(
            "Nominal Aktual Ads (Rp)", min_value=None, format="accounting"
        ),
    }


def initialize_non_ads_lainnya_data_session():
    """
    Menginisialisasi DataFrame di st.session_state untuk brand tertentu jika belum ada.
    """
    session_key = "df_non_ads_lainnya"

    if session_key not in st.session_state:
        # Buat DataFrame default dengan kolom yang dibutuhkan
        data = {
            "Tanggal": pd.Series(get_yesterday_in_jakarta(), index=range(1)),
            "Nama Project": ["Enzhico"],
            "Keterangan": ["Talent"],
            "Nominal Aktual Non Ads": [0.0],
        }
        st.session_state[session_key] = pd.DataFrame(data)


def get_non_ads_lainnya_column_config():
    """
    Mengembalikan konfigurasi kolom yang konsisten untuk st.data_editor.
    """
    return {
        "Tanggal": st.column_config.DateColumn(
            "Tanggal",
            min_value=pd.Timestamp(2023, 1, 1),
            default=get_yesterday_in_jakarta(),
            format="YYYY-MM-DD",
            required=True,
        ),
        "Nama Project": st.column_config.SelectboxColumn(
            "Nama Project", options=PROJECT_NAME_LIST, required=True
        ),
        "Keterangan": st.column_config.TextColumn("Keterangan", required=True),
        "Nominal Aktual Non Ads": st.column_config.NumberColumn(
            "Nominal Aktual Non Ads (Rp)",
            min_value=0.0,
            format="accounting",
            required=True,
        ),
    }


def initialize_non_ads_data_session(project_name, marketplace_list, store_list):
    """
    Menginisialisasi DataFrame di st.session_state untuk project tertentu jika belum ada.

    Args:
        project_name (str): Nama project untuk kunci di session_state.
        marketplace__list (list): Daftar nama marketplace yang akan diisi ke DataFrame.
        store_list (list): Daftar nama toko yang akan diisi ke DataFrame.
    """
    session_key = f"df_{project_name}_ads"

    if session_key not in st.session_state:
        # Buat DataFrame default dengan kolom yang dibutuhkan
        data = {
            "Tanggal": pd.Series(
                get_yesterday_in_jakarta(),
                index=range(len(store_list)),
            ),
            "Marketplace": marketplace_list,
            "Nama Toko": store_list,
            "Nominal Aktual Non Ads": None,
        }
        st.session_state[session_key] = pd.DataFrame(data)


def get_non_ads_column_config(store_list):
    """
    Mengembalikan konfigurasi kolom yang konsisten untuk st.data_editor.
    """

    return {
        "Tanggal": st.column_config.DateColumn(
            "Tanggal",
            min_value=pd.Timestamp(2023, 1, 1),
            format="YYYY-MM-DD",
            required=True,
        ),
        "Marketplace": st.column_config.SelectboxColumn(
            "Marketplace",
            options=MARKETPLACE_LIST,
            required=True,
        ),
        "Nama Toko": st.column_config.SelectboxColumn(
            "Nama Toko",
            options=store_list,
            required=True,
        ),
        "Nominal Aktual Non Ads": st.column_config.NumberColumn(
            "Nominal Aktual Non Ads (Rp)",
            min_value=0.0,
            format="accounting",
        ),
    }


def initialize__ads_reg_data_session(branch_name):
    """
    Menginisialisasi DataFrame di st.session_state untuk brand tertentu jika belum ada.
    """
    session_key = f"df_{branch_name}_reg_ads"

    if session_key not in st.session_state:
        # Buat DataFrame default dengan kolom yang dibutuhkan
        data = {
            "Tanggal": pd.Series(get_yesterday_in_jakarta(), index=range(3)),
            "Akun": AKUN_REGULAR,
            "Nominal Aktual Ads": None,
        }
        st.session_state[session_key] = pd.DataFrame(data)


def get_ads_reg_column_config():
    """
    Mengembalikan konfigurasi kolom yang konsisten untuk st.data_editor.
    """
    return {
        "Tanggal": st.column_config.DateColumn(
            "Tanggal",
            min_value=pd.Timestamp(2023, 1, 1),
            format="YYYY-MM-DD",
            required=True,
        ),
        "Akun": st.column_config.SelectboxColumn("Akun", options=AKUN_REGULAR),
        "Nominal Aktual Ads": st.column_config.NumberColumn(
            "Nominal Aktual Ads (Rp)",
            min_value=0.0,
            format="accounting",
        ),
    }


# -- STOCK --
def initialize_stock_data_session():
    """
    Menginisialisasi DataFrame di st.session_state untuk brand tertentu jika belum ada.

    """
    session_key = "df_stock"

    if session_key not in st.session_state:
        # Buat DataFrame default dengan kolom yang dibutuhkan
        data = {
            "Tanggal": pd.Series(get_yesterday_in_jakarta(), index=range(5)),
            "Marketplace": None,
            "Nama Toko": None,
            "Nominal Budget Ads": [0.0] * 5,
            "Nominal Aktual Ads": None,
        }
        st.session_state[session_key] = pd.DataFrame(data)


def get_stock_column_config(store_list):
    """
    Mengembalikan konfigurasi kolom yang konsisten untuk st.data_editor.
    """

    return {
        "Tanggal": st.column_config.DateColumn(
            "Tanggal",
            min_value=pd.Timestamp(2023, 1, 1),
            format="YYYY-MM-DD",
            required=True,
        ),
        "Marketplace": st.column_config.SelectboxColumn(
            "Marketplace",
            options=MARKETPLACE_LIST,
            required=True,
        ),
        "Nama Toko": st.column_config.SelectboxColumn(
            "Nama Toko",
            options=store_list,
            required=True,
        ),
        "Nominal Budget Ads": st.column_config.NumberColumn(
            "Nominal Budget Ads (Rp)",
            min_value=0.0,
            format="accounting",
            required=True,
        ),
        "Nominal Aktual Ads": st.column_config.NumberColumn(
            "Nominal Aktual Ads (Rp)",
            min_value=0.0,
            format="accounting",
        ),
    }


# ADVERTISER DATA
def initialize_marketplace_data_session(project_name, marketplace_list, store_list):
    """
    Menginisialisasi DataFrame di st.session_state untuk brand tertentu jika belum ada.
    """
    session_key = f"df_{project_name}_marketplace"

    if session_key not in st.session_state:
        # Buat DataFrame default dengan kolom yang dibutuhkan
        data = {
            "Tanggal": pd.Series(
                get_yesterday_in_jakarta(),
                index=range(len(store_list)),
            ),
            "Marketplace": marketplace_list,
            "Nama Toko": store_list,
            "Spend": [0.0] * len(store_list),
            "Konversi": [0] * len(store_list),
            "Produk Terjual": [0] * len(store_list),
            "Gross Revenue": [0.0] * len(store_list),
            "CTR": [0.0] * len(store_list),
        }
        st.session_state[session_key] = pd.DataFrame(data)


def get_marketplace_column_config(store_list):
    """
    Mengembalikan konfigurasi kolom yang konsisten untuk st.data_editor.
    """

    return {
        "Tanggal": st.column_config.DateColumn(
            "Tanggal",
            min_value=pd.Timestamp(2023, 1, 1),
            format="YYYY-MM-DD",
            required=True,
        ),
        "Marketplace": st.column_config.SelectboxColumn(
            "Marketplace",
            options=MARKETPLACE_LIST,
            required=True,
        ),
        "Nama Toko": st.column_config.SelectboxColumn(
            "Nama Toko",
            options=store_list,
            required=True,
        ),
        "Spend": st.column_config.NumberColumn(
            "Spend (Rp)", min_value=None, format="accounting"
        ),
        "Konversi": st.column_config.NumberColumn(
            "Konversi", min_value=None, step=1, format="%d"
        ),
        "Produk Terjual": st.column_config.NumberColumn(
            "Produk Terjual",
            min_value=None,
            step=1,
            format="%d",
        ),
        "Gross Revenue": st.column_config.NumberColumn(
            "Gross Revenue (Rp)",
            min_value=None,
            format="accounting",
        ),
        "CTR": st.column_config.NumberColumn(
            "CTR",
            min_value=None,
            format="%.2f",
        ),
    }


def initialize_cpas_data_session(project_name, store_list, akun_list):
    """
    Menginisialisasi DataFrame di st.session_state untuk store_list tertentu jika belum ada.

    Args:
        project_name (str): Nama store_list untuk kunci di session_state.
        marketplace__list (list): Daftar nama marketplace yang akan diisi ke DataFrame.
        akun (list): Daftar nama toko yang akan diisi ke DataFrame.
    """
    session_key = f"df_{project_name}_cpas"

    if session_key not in st.session_state:
        # Buat DataFrame default dengan kolom yang dibutuhkan
        data = {
            "Tanggal": pd.Series(
                get_yesterday_in_jakarta(),
                index=range(len(akun_list)),
            ),
            "Nama Toko": store_list,
            "Akun": akun_list,
            "Spend": [0.0] * len(akun_list),
            "Konversi": [0] * len(akun_list),
            "Gross Revenue": [0.0] * len(akun_list),
        }
        st.session_state[session_key] = pd.DataFrame(data)


def get_cpas_column_config(store_list, akun_list):
    """
    Mengembalikan konfigurasi kolom yang konsisten untuk st.data_editor.
    """

    return {
        "Tanggal": st.column_config.DateColumn(
            "Tanggal",
            min_value=pd.Timestamp(2023, 1, 1),
            format="YYYY-MM-DD",
            required=True,
        ),
        "Nama Toko": st.column_config.SelectboxColumn(
            "Nama Toko",
            options=store_list,
            required=True,
        ),
        "Akun": st.column_config.SelectboxColumn(
            "Akun",
            options=akun_list,
            required=True,
        ),
        "Spend": st.column_config.NumberColumn(
            "Spend (Rp)", min_value=None, format="accounting"
        ),
        "Konversi": st.column_config.NumberColumn(
            "Konversi", min_value=None, step=1, format="%d"
        ),
        "Gross Revenue": st.column_config.NumberColumn(
            "Gross Revenue (Rp)",
            min_value=None,
            format="accounting",
        ),
    }


# DASHBOARD ADMIN
def generate_excel_bytes(df, group_cols, value_col, aggfunc="sum"):
    """
    Membuat report Excel dari hasil groupby dan mengembalikannya sebagai
    objek bytes untuk di-download.
    """
    if not group_cols:
        raise ValueError("Parameter 'group_cols' tidak boleh kosong.")

    brand_col = group_cols[0]
    grouped = (
        df.groupby(group_cols, as_index=False)[value_col]
        .agg(aggfunc)
        .sort_values(by=brand_col)
    )
    grand_total = grouped[value_col].sum()

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    header = group_cols + [value_col.capitalize()]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row_idx = 2
    for brand_name, brand_df in grouped.groupby(brand_col):
        start_merge_row = current_row_idx
        for _, data_row in brand_df.iterrows():
            ws.append(list(data_row))
            current_row_idx += 1

        subtotal_value = brand_df[value_col].sum()
        subtotal_row_data = (
            [f"Subtotal {brand_name}"] + [""] * (len(group_cols) - 1) + [subtotal_value]
        )
        ws.append(subtotal_row_data)
        for cell in ws[current_row_idx]:
            cell.font = Font(bold=True)
        current_row_idx += 1

        end_merge_row = current_row_idx - 2
        if start_merge_row < end_merge_row:
            ws.merge_cells(
                start_row=start_merge_row,
                start_column=1,
                end_row=end_merge_row,
                end_column=1,
            )
            ws.cell(start_merge_row, 1).alignment = Alignment(
                horizontal="left", vertical="center"
            )

    grand_total_row_data = (
        ["Grand Total"] + [""] * (len(group_cols) - 1) + [grand_total]
    )
    ws.append(grand_total_row_data)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, size=12)

    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # --- PERUBAHAN UTAMA ADA DI SINI ---
    # Simpan workbook ke dalam buffer memori
    buffer = io.BytesIO()
    wb.save(buffer)
    # Pindahkan "cursor" buffer ke awal
    buffer.seek(0)
    return buffer


def expand_sku(sku: str):
    """
    Expand SKU bundling sesuai aturan:
    - Jika tidak ada " + ", return [sku]
    - Jika ada "_", pecah jadi beberapa produk
    """
    if " + " not in sku:
        return [sku]

    parts = sku.split(" + ")
    prefix = parts[0].split("-")[0]  # ambil brand prefix (3 huruf pertama sebelum '-')

    expanded = [parts[0]]  # produk pertama utuh
    for p in parts[1:]:
        expanded.append(prefix + "-" + p)

    return expanded


def create_visual_report(report_df, original_df):
    """
    Membuat laporan Excel yang sangat detail dan mudah dibaca, dengan grouping
    dan subtotal/grand total per hari, marketplace, dan toko untuk Jumlah PCS
    serta Total Resi Unik.
    """
    if report_df.empty or original_df.empty:
        wb = Workbook()
        ws = wb.active
        ws.title = "Laporan Marketplace"
        ws["A1"] = "Tidak ada data untuk ditampilkan."
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # --- 1. PRA-PEMROSESAN DATA DENGAN PANDAS ---
    # Salin DataFrame agar tidak mengubah data asli
    df = report_df.copy()
    df_orig = original_df.copy()

    # Pastikan ada kolom 'tanggal' berformat date
    df["tanggal"] = pd.to_datetime(df["timestamp_input_data"]).dt.date
    df_orig["tanggal"] = pd.to_datetime(df_orig["timestamp_input_data"]).dt.date

    # a. Agregasi level paling detail: Jumlah PCS per SKU per hari
    df_details = df.groupby(
        ["tanggal", "nama_marketplace", "nama_toko", "sku"], as_index=False
    ).agg(jumlah_pcs=("jumlah_item", "sum"))

    # b. Agregasi Resi Unik di semua level yang dibutuhkan
    df_resi_toko = df_orig.groupby(
        ["tanggal", "nama_marketplace", "nama_toko"], as_index=False
    ).agg(total_resi_unik_toko=("no_resi", "nunique"))

    df_resi_marketplace = df_orig.groupby(
        ["tanggal", "nama_marketplace"], as_index=False
    ).agg(total_resi_unik_marketplace=("no_resi", "nunique"))

    df_resi_hari = df_orig.groupby(["tanggal"], as_index=False).agg(
        total_resi_unik_hari=("no_resi", "nunique")
    )

    # c. Gabungkan semua data menjadi satu DataFrame utama
    df_final = pd.merge(
        df_details,
        df_resi_toko,
        on=["tanggal", "nama_marketplace", "nama_toko"],
        how="left",
    )
    df_final = pd.merge(
        df_final, df_resi_marketplace, on=["tanggal", "nama_marketplace"], how="left"
    )
    df_final = pd.merge(df_final, df_resi_hari, on=["tanggal"], how="left")
    df_final.fillna(0, inplace=True)  # Ganti NaN dengan 0

    # d. Urutkan DataFrame (KRUSIAL untuk logika loop di bawah)
    df_sorted = df_final.sort_values(
        by=["tanggal", "nama_marketplace", "nama_toko", "sku"]
    ).reset_index(drop=True)

    # --- 2. PERSIAPAN FILE EXCEL ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Rinci Marketplace"

    # Styles
    bold_font = Font(bold=True, name="Calibri")
    italic_bold_font = Font(bold=True, italic=True, name="Calibri")
    header_fill = PatternFill(
        start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
    )
    toko_fill = PatternFill(start_color="FDF2CC", end_color="FDF2CC", fill_type="solid")
    mp_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    hari_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    grand_total_fill = PatternFill(
        start_color="A9D08E", end_color="A9D08E", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header
    headers = [
        "Tanggal",
        "Marketplace",
        "Nama Toko",
        "SKU",
        "Jumlah PCS",
        "Resi Unik Toko",
        "Resi Unik Marketplace",
        "Resi Unik per Tanggal",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = thin_border

    # --- 3. LOOPING & PENULISAN KE EXCEL ---
    # Variabel pelacak status & akumulator total
    prev = {"toko": None, "mp": None, "tgl": None}
    total = {"toko": 0, "mp": 0, "tgl": 0}

    # Fungsi bantuan untuk menulis baris subtotal
    def write_subtotal_row(label, pcs_total, resi_total, col_idx_resi, fill, font):
        row_data = [""] * len(headers)
        row_data[2] = label  # Kolom C untuk label
        row_data[4] = pcs_total  # Kolom E untuk Jml PCS
        row_data[col_idx_resi] = resi_total  # Kolom F/G/H untuk Resi
        ws.append(row_data)
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.font = font
            cell.border = thin_border

    for idx, row in df_sorted.iterrows():
        # Cek perubahan grup dari level terbesar (tanggal) ke terkecil (toko)
        is_new_day = prev["tgl"] and row["tanggal"] != prev["tgl"]
        is_new_mp = prev["mp"] and row["nama_marketplace"] != prev["mp"]
        is_new_toko = prev["toko"] and row["nama_toko"] != prev["toko"]

        # Jika hari baru, tutup semua subtotal dari hari sebelumnya
        if is_new_day:
            prev_row = df_sorted.loc[idx - 1]
            write_subtotal_row(
                f"Subtotal Toko: {prev['toko']}",
                total["toko"],
                prev_row["total_resi_unik_toko"],
                5,
                toko_fill,
                italic_bold_font,
            )
            write_subtotal_row(
                f"Subtotal Marketplace: {prev['mp']}",
                total["mp"],
                prev_row["total_resi_unik_marketplace"],
                6,
                mp_fill,
                bold_font,
            )
            write_subtotal_row(
                f"GRAND TOTAL TANGGAL: {prev['tgl']}",
                total["tgl"],
                prev_row["total_resi_unik_hari"],
                7,
                hari_fill,
                bold_font,
            )
            ws.append([])  # Baris kosong pemisah hari
            total = {"toko": 0, "mp": 0, "tgl": 0}

        # Jika marketplace baru (di hari yang sama)
        elif is_new_mp:
            prev_row = df_sorted.loc[idx - 1]
            write_subtotal_row(
                f"Subtotal Toko: {prev['toko']}",
                total["toko"],
                prev_row["total_resi_unik_toko"],
                5,
                toko_fill,
                italic_bold_font,
            )
            write_subtotal_row(
                f"Subtotal Marketplace: {prev['mp']}",
                total["mp"],
                prev_row["total_resi_unik_marketplace"],
                6,
                mp_fill,
                bold_font,
            )
            ws.append([])  # Baris kosong pemisah marketplace
            total["toko"], total["mp"] = 0, 0

        # Jika hanya toko yang baru (di marketplace & hari yang sama)
        elif is_new_toko:
            prev_row = df_sorted.loc[idx - 1]
            write_subtotal_row(
                f"Subtotal Toko: {prev['toko']}",
                total["toko"],
                prev_row["total_resi_unik_toko"],
                5,
                toko_fill,
                italic_bold_font,
            )
            total["toko"] = 0

        # Tulis baris data detail
        ws.append(
            [
                row["tanggal"] if row["tanggal"] != prev["tgl"] else "",
                row["nama_marketplace"]
                if row["nama_marketplace"] != prev["mp"]
                else "",
                row["nama_toko"] if row["nama_toko"] != prev["toko"] else "",
                row["sku"],
                row["jumlah_pcs"],
                "",
                "",
                "",  # Resi unik hanya diisi di subtotal
            ]
        )
        for cell in ws[ws.max_row]:
            cell.border = thin_border

        # Akumulasi total
        total["toko"] += row["jumlah_pcs"]
        total["mp"] += row["jumlah_pcs"]
        total["tgl"] += row["jumlah_pcs"]

        # Update pelacak
        prev = {
            "toko": row["nama_toko"],
            "mp": row["nama_marketplace"],
            "tgl": row["tanggal"],
        }

    # --- Tulis Subtotal TERAKHIR setelah loop selesai ---
    if prev["tgl"]:
        last_row = df_sorted.iloc[-1]
        write_subtotal_row(
            f"Subtotal Toko: {prev['toko']}",
            total["toko"],
            last_row["total_resi_unik_toko"],
            5,
            toko_fill,
            italic_bold_font,
        )
        write_subtotal_row(
            f"Subtotal Marketplace: {prev['mp']}",
            total["mp"],
            last_row["total_resi_unik_marketplace"],
            6,
            mp_fill,
            bold_font,
        )
        write_subtotal_row(
            f"GRAND TOTAL TANGGAL: {prev['tgl']}",
            total["tgl"],
            last_row["total_resi_unik_hari"],
            7,
            hari_fill,
            bold_font,
        )

    # --- Tulis Grand Total Keseluruhan ---
    ws.append([])
    grand_total_pcs = df_sorted["jumlah_pcs"].sum()
    grand_total_resi = df_orig["no_resi"].nunique()
    write_subtotal_row(
        "GRAND TOTAL KESELURUHAN",
        grand_total_pcs,
        grand_total_resi,
        7,
        grand_total_fill,
        bold_font,
    )

    # --- 4. FORMATTING AKHIR ---
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 2

    # Simpan ke memori
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def parse_bundle_pcs(row):
    sku = row.iloc[0]  # kolom pertama (sku)
    qty = row.iloc[1]  # kolom kedua (jumlah_item)

    match = re.search(
        r"-(\d+)-(PCS|BOX|PAK|PACK|BTL|LTR|KG)$", sku, flags=re.IGNORECASE
    )
    if match:
        jumlah_bundle = int(match.group(1))
        qty = qty * jumlah_bundle
        sku = re.sub(
            r"-(\d+)-(PCS|BOX|PAK|PACK|BTL|LTR|KG)$", "", sku, flags=re.IGNORECASE
        )
    return pd.Series([sku, qty])


def get_quarter_months(month: int):
    """Helper: menentukan bulan dalam kuartal"""
    if month in [1, 2, 3]:
        return ["January", "February", "March"], 1
    elif month in [4, 5, 6]:
        return ["April", "May", "June"], 2
    elif month in [7, 8, 9]:
        return ["July", "August", "September"], 3
    else:
        return ["October", "November", "December"], 4


def transform_budget_to_report(df_raw_budget: pd.DataFrame) -> pd.DataFrame:
    """
    Mengubah DataFrame budget mentah (format panjang) menjadi format laporan (format lebar)
    menggunakan pivot.
    """
    if df_raw_budget.empty:
        return pd.DataFrame()

    df = df_raw_budget.drop(columns=["id", "created_at", "project_id"], errors="ignore")

    pivot_df = df.pivot_table(
        index=[
            "project_name",
            "parameter_name",
            "target_rasio_persen",
            "tahun",
            "kuartal",
        ],
        columns="bulan",
        values="target_bulanan_rp",
        aggfunc="sum",
    ).reset_index()

    pivot_df.columns.name = None

    bulan_order = [
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ]
    existing_months = [month for month in bulan_order if month in pivot_df.columns]

    if existing_months:
        pivot_df["Target Kuartal"] = (pivot_df[existing_months[0]] * 3).round(-1)
    else:
        pivot_df["Target Kuartal"] = 0

    final_columns_order = (
        ["project_name", "parameter_name", "target_rasio_persen", "Target Kuartal"]
        + existing_months
        + ["tahun", "kuartal"]
    )

    report_df = pivot_df[final_columns_order]

    report_df = report_df.rename(
        columns={
            "project_name": "Project",
            "parameter_name": "Parameter",
            "target_rasio_persen": "Target Rasio",
            "tahun": "Tahun",
            "kuartal": "Kuartal",
        }
    )

    return report_df


# ADV REGULAR
def initialize_adv_reg_data_session(platform, project_list, produk_list):
    """
    Menginisialisasi DataFrame di st.session_state untuk brand tertentu jika belum ada.
    """
    session_key = f"df_{platform}_adv_reg"

    if session_key not in st.session_state:
        # Buat DataFrame default dengan kolom yang dibutuhkan
        data = {
            "Tanggal": pd.Series(
                get_yesterday_in_jakarta(),
                index=range(len(produk_list)),
            ),
            "Produk": project_list,
            "Spend": [0.0] * len(produk_list),
            "Reach": [0.0] * len(produk_list),
            "CTR": [0.0] * len(produk_list),
            "Thruplays": [0.0] * len(produk_list),
            "Lead": [0] * len(produk_list),
        }
        st.session_state[session_key] = pd.DataFrame(data)


def get_adv_reg_column_config(store_list):
    """
    Mengembalikan konfigurasi kolom yang konsisten untuk st.data_editor.
    """

    return {
        "Tanggal": st.column_config.DateColumn(
            "Tanggal",
            min_value=pd.Timestamp(2023, 1, 1),
            format="YYYY-MM-DD",
            required=True,
        ),
        "Produk": st.column_config.SelectboxColumn(
            "Produk",
            options=MARKETPLACE_LIST,
            required=True,
        ),
        "Nama Toko": st.column_config.SelectboxColumn(
            "Nama Toko",
            options=store_list,
            required=True,
        ),
        "Akrual Basis": st.column_config.NumberColumn(
            "Akrual Basis (Rp)",
            min_value=0.0,
            format="accounting",
            required=True,
        ),
        "Cash Basis": st.column_config.NumberColumn(
            "Cash Basis (Rp)",
            min_value=0.0,
            format="accounting",
        ),
        "Bukti": st.column_config.TextColumn("Bukti"),
        "Akun Bank": st.column_config.TextColumn("Akun Bank"),
    }


@st.cache_data(ttl=30)
def fetch_data(_conn):
    """Mengambil semua data, diurutkan dengan channel."""
    query = "SELECT * FROM advertiser_cs_regular ORDER BY performance_date DESC, product_name, channel;"
    try:
        df = pd.read_sql(query, _conn)
        df["performance_date"] = pd.to_datetime(df["performance_date"]).dt.date
        return df
    except Exception as e:
        st.error(f"Gagal mengambil data: {e}")
        return pd.DataFrame()


def create_daily_template(conn, template_date, product_channel_list):
    """Memasukkan baris template dengan menyertakan channel."""
    with conn.cursor() as cur:
        for product, channel in product_channel_list:
            query = """
                INSERT INTO advertiser_cs_regular (performance_date, product_name, channel)
                VALUES (%s, %s, %s)
                ON CONFLICT (performance_date, product_name, channel) DO NOTHING;
            """
            cur.execute(query, (template_date, product, channel))
    conn.commit()


def process_changes(conn, original_df, changes):
    """Memproses perubahan dengan primary key tiga bagian."""
    with conn.cursor() as cur:
        # Proses Hapus
        for index in changes["deleted_rows"]:
            row = original_df.iloc[index]
            cur.execute(
                "DELETE FROM advertiser_cs_regular WHERE performance_date = %s AND product_name = %s AND channel = %s;",
                (row["performance_date"], row["product_name"], row["channel"]),
            )

        # Proses Tambah (UPSERT)
        for new_row in changes["added_rows"]:
            new_row.pop("_index", None)
            if not all(
                new_row.get(k) for k in ["performance_date", "product_name", "channel"]
            ):
                continue

            cols = [sql.Identifier(k) for k, v in new_row.items() if v is not None]
            if not cols:
                continue
            vals = [v for v in new_row.values() if v is not None]
            update_assignments = sql.SQL(", ").join(
                sql.SQL("{col} = EXCLUDED.{col}").format(col=col)
                for col in cols
                if col.string not in ["performance_date", "product_name", "channel"]
            )

            query = sql.SQL("""
                INSERT INTO advertiser_cs_regular ({cols}) VALUES ({vals})
                ON CONFLICT (performance_date, product_name, channel) DO UPDATE SET {updates};
            """).format(
                cols=sql.SQL(", ").join(cols),
                vals=sql.SQL(", ").join(sql.Placeholder() * len(vals)),
                updates=update_assignments,
            )
            cur.execute(query, vals)

        # Proses Edit
        for index, updates in changes["edited_rows"].items():
            updates.pop("_index", None)
            if not updates:
                continue

            original_row = original_df.iloc[index]
            set_clauses = [
                sql.SQL("{} = %s").format(sql.Identifier(col)) for col in updates.keys()
            ]
            query = sql.SQL(
                "UPDATE advertiser_cs_regular SET {} WHERE performance_date = %s AND product_name = %s AND channel = %s;"
            ).format(sql.SQL(", ").join(set_clauses))

            params = list(updates.values()) + [
                original_row["performance_date"],
                original_row["product_name"],
                original_row["channel"],
            ]
            cur.execute(query, params)
    conn.commit()
